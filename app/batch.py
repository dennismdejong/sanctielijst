"""Batch screening module: input parsing (CSV/XLSX), SQLite job store and background processing.

The module is decoupled from the search implementation: callers inject a
``search_fn`` callable with signature
``search_fn(naam, geboortejaar=None, nationaliteit=None, geboorteplaats=None, type=None)``
that returns a JSON-serialisable list of match results. ``main.py`` wraps
``run_search`` in such a callable; tests inject a fake.
"""

import csv
import json
import os
import sqlite3
import uuid
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

FIELDS = ["naam", "geboortejaar", "nationaliteit", "geboorteplaats", "type"]
DEFAULT_ROW_LIMIT = 5000
MAX_BATCH_BYTES = 50 * 1024 * 1024

SCHEMA = """
CREATE TABLE IF NOT EXISTS batch_jobs (
  id TEXT PRIMARY KEY,
  status TEXT NOT NULL,
  created_at TEXT NOT NULL,
  finished_at TEXT,
  progress INTEGER NOT NULL DEFAULT 0,
  total INTEGER NOT NULL DEFAULT 0,
  errors TEXT NOT NULL DEFAULT '[]',
  error_text TEXT
)
"""

RESULTS_SCHEMA = """
CREATE TABLE IF NOT EXISTS batch_results (
  batch_id TEXT NOT NULL,
  row_index INTEGER NOT NULL,
  row_json TEXT NOT NULL,
  matches_json TEXT
)
"""


class BatchInputError(Exception):
    """Parse-level input error; ``status_code`` maps to an HTTP status (default 400)."""

    status_code = 400


class RowLimitExceeded(BatchInputError):
    """More input rows than ``row_limit``; maps to HTTP 413."""

    status_code = 413


def default_batch_db() -> Path:
    return Path(os.environ.get("BATCH_DB") or str(Path("data") / "batch.sqlite"))


def _open(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def init_batch_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with _open(db_path) as conn:
        conn.execute(SCHEMA)
        conn.execute(RESULTS_SCHEMA)
        conn.commit()


def _map_columns(header: list | tuple) -> dict[str, int]:
    mapping = {}
    for index, value in enumerate(header):
        key = str(value).strip().lower() if value is not None else ""
        key = {"name": "naam"}.get(key, key)
        if key in FIELDS and key not in mapping:
            mapping[key] = index
    return mapping


def _normalise_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalise_birth_year(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _build_rows(mapping: dict[str, int], raw_rows: list[tuple[int, list | tuple]]) -> tuple[list[dict], list[dict]]:
    rows = []
    errors = []
    for row_index, values in raw_rows:
        row = {field: None for field in FIELDS}
        for field, index in mapping.items():
            if index < len(values):
                value = values[index]
                if field == "geboortejaar":
                    birth = _normalise_birth_year(value)
                    if birth is None and value is not None and str(value).strip():
                        errors.append({"row_index": row_index, "error": "Ongeldig geboortejaar"})
                    row[field] = birth
                else:
                    row[field] = _normalise_value(value)
        if not row["naam"]:
            errors.append({"row_index": row_index, "error": "Ontbrekende naam"})
            continue
        rows.append(row)
    if not rows:
        raise BatchInputError("Geen geldige regels gevonden")
    return rows, errors


def _detect_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,").delimiter
    except csv.Error:
        return ";"


def _decode_text(content: bytes) -> str:
    """Decode input bytes defensively: UTF-8 (with BOM) first, then CP1252/Latin-1."""
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    try:
        return content.decode("cp1252")
    except UnicodeDecodeError:
        pass
    return content.decode("latin-1")


def _parse_csv(content: bytes, row_limit: int) -> tuple[list[dict], list[dict]]:
    text = _decode_text(content)
    if not text.strip():
        raise BatchInputError("Bestand is leeg")
    delimiter = _detect_delimiter(text.strip())
    reader = csv.reader(StringIO(text, newline=""), delimiter=delimiter)
    mapping = {}
    raw_rows = []
    for values in reader:
        if not values or all(str(value).strip() == "" for value in values):
            continue
        if not mapping:
            mapping = _map_columns(values)
            if "naam" not in mapping:
                raise BatchInputError("Naam-kolom ontbreekt")
            continue
        raw_rows.append((len(raw_rows) + 2, values))
        if len(raw_rows) > row_limit:
            raise RowLimitExceeded(f"Bestand bevat meer dan {row_limit} regels")
    if not mapping:
        raise BatchInputError("Bestand is leeg")
    return _build_rows(mapping, raw_rows)


def _parse_xlsx(content: bytes, row_limit: int) -> tuple[list[dict], list[dict]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True)
    except Exception as exc:
        raise BatchInputError("Ongeldig Excel-bestand") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if header is None or all(value is None for value in header):
            raise BatchInputError("Bestand is leeg")
        mapping = _map_columns(header)
        if "naam" not in mapping:
            raise BatchInputError("Naam-kolom ontbreekt")
        raw_rows = []
        for row_index, values in enumerate(iterator, start=2):
            if values is None or all(value is None or str(value).strip() == "" for value in values):
                continue
            raw_rows.append((row_index, values))
            if len(raw_rows) > row_limit:
                raise RowLimitExceeded(f"Bestand bevat meer dan {row_limit} regels")
        return _build_rows(mapping, raw_rows)
    except BatchInputError:
        raise
    except Exception as exc:
        raise BatchInputError("Ongeldig Excel-bestand") from exc
    finally:
        workbook.close()


def parse_input(filename: str, content: bytes, row_limit: int = DEFAULT_ROW_LIMIT) -> tuple[list[dict], list[dict]]:
    """Parse an uploaded batch file into rows plus per-row errors.

    ``filename`` selects the format: ``.xlsx`` is read with openpyxl, everything
    else with the stdlib ``csv`` module. Header column names are matched
    case-insensitively against the row schema. An empty file, a missing name
    column, or more than ``row_limit`` rows raise ``BatchInputError`` (413 for
    the row limit); rows without a name are reported per-row and skipped.
    """
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content, row_limit)
    return _parse_csv(content, row_limit)


def create_job(db_path: Path, filename: str, rows: list[dict], errors: list[dict] | None = None) -> str:
    # ``filename`` is part of the public interface but the mandated schema does
    # not persist it; it is reserved for display/logging by the API layer.
    # ``errors`` (per-row parse errors) is persisted as JSON in the errors column.
    init_batch_db(db_path)
    job_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    errors_json = json.dumps(errors or [], ensure_ascii=False)
    with _open(db_path) as conn:
        conn.execute(
            "INSERT INTO batch_jobs (id, status, created_at, progress, total, errors) VALUES (?, ?, ?, 0, ?, ?)",
            (job_id, "pending", now, len(rows), errors_json),
        )
        conn.executemany(
            "INSERT INTO batch_results (batch_id, row_index, row_json, matches_json) VALUES (?, ?, ?, NULL)",
            [(job_id, index, json.dumps(row, ensure_ascii=False)) for index, row in enumerate(rows)],
        )
        conn.commit()
    return job_id


def get_job(db_path: Path, job_id: str) -> dict | None:
    init_batch_db(db_path)
    with _open(db_path) as conn:
        row = conn.execute("SELECT * FROM batch_jobs WHERE id = ?", (job_id,)).fetchone()
    if row is None:
        return None
    job = dict(row)
    job["errors"] = json.loads(job["errors"]) if job["errors"] else []
    return job


def get_results(db_path: Path, job_id: str) -> list[dict]:
    init_batch_db(db_path)
    with _open(db_path) as conn:
        rows = conn.execute(
            "SELECT row_index, row_json, matches_json FROM batch_results WHERE batch_id = ? ORDER BY row_index",
            (job_id,),
        ).fetchall()
    results = []
    for row in rows:
        results.append({
            "row_index": row["row_index"],
            "row": json.loads(row["row_json"]),
            "matches": json.loads(row["matches_json"]) if row["matches_json"] is not None else None,
        })
    return results


def mark_stale_jobs(
    db_path: Path,
    stale_statuses: tuple[str, ...] = ("pending", "running"),
    error_text: str = "Onderbroken door herstart",
) -> int:
    """Mark orphaned batch jobs as ``error`` after a restart.

    ``process_job`` runs in a daemon thread with no persistence hook; if the
    process dies mid-batch the job would otherwise stay ``pending``/``running``
    forever and the report endpoints would permanently 404. Called from
    ``create_app`` at startup; a no-op when the batch DB does not exist yet.
    Returns the number of jobs updated.
    """
    if not db_path.exists():
        return 0
    placeholders = ",".join("?" for _ in stale_statuses)
    with _open(db_path) as conn:
        cursor = conn.execute(
            f"UPDATE batch_jobs SET status = 'error', finished_at = ?, error_text = ? WHERE status IN ({placeholders})",
            (datetime.now(timezone.utc).isoformat(), error_text, *stale_statuses),
        )
        conn.commit()
        return cursor.rowcount


def process_job(db_path: Path, job_id: str, search_fn, row_limit: int = DEFAULT_ROW_LIMIT) -> None:
    """Process every pending result row through ``search_fn`` and store the matches.

    Input rows are written to ``batch_results`` by ``create_job`` (with NULL
    ``matches_json``); this updates them in order and tracks ``progress``.
    """
    init_batch_db(db_path)
    with _open(db_path) as conn:
        job = conn.execute("SELECT status, total FROM batch_jobs WHERE id = ?", (job_id,)).fetchone()
        if job is None:
            raise KeyError(job_id)
        if job["total"] > row_limit:
            raise RowLimitExceeded(f"Bestand bevat meer dan {row_limit} regels")
        pending = conn.execute(
            "SELECT row_index, row_json FROM batch_results WHERE batch_id = ? AND matches_json IS NULL ORDER BY row_index",
            (job_id,),
        ).fetchall()
        if not pending:
            return
        base_progress = conn.execute(
            "SELECT COUNT(*) FROM batch_results WHERE batch_id = ? AND matches_json IS NOT NULL",
            (job_id,),
        ).fetchone()[0]
        conn.execute("UPDATE batch_jobs SET status = 'running' WHERE id = ?", (job_id,))
        conn.commit()
        try:
            for index, item in enumerate(pending):
                row = json.loads(item["row_json"])
                matches = search_fn(
                    row["naam"],
                    geboortejaar=row.get("geboortejaar"),
                    nationaliteit=row.get("nationaliteit"),
                    geboorteplaats=row.get("geboorteplaats"),
                    type=row.get("type"),
                )
                conn.execute(
                    "UPDATE batch_results SET matches_json = ? WHERE batch_id = ? AND row_index = ?",
                    (json.dumps(matches, ensure_ascii=False), job_id, item["row_index"]),
                )
                conn.execute("UPDATE batch_jobs SET progress = ? WHERE id = ?", (base_progress + index + 1, job_id))
                conn.commit()
            conn.execute(
                "UPDATE batch_jobs SET status = 'done', finished_at = ?, progress = ? WHERE id = ?",
                (datetime.now(timezone.utc).isoformat(), job["total"], job_id),
            )
            conn.commit()
        except Exception as exc:
            conn.execute(
                "UPDATE batch_jobs SET status = 'error', finished_at = ?, error_text = ? WHERE id = ?",
                (datetime.now(timezone.utc).isoformat(), f"Verwerking mislukt: {exc!r}", job_id),
            )
            conn.commit()
            raise
