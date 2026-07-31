import csv
import re
import zlib
from io import BytesIO, StringIO

import openpyxl
from reportlab.pdfbase.pdfutils import asciiBase85Decode

from app.export import render_search_pdf, _escape, _export_rows, render_search_csv, render_search_xlsx


def _payload(**over):
    payload = {
        "query": {"name": "JORGE FERNANDEZ", "birth_year": None, "nationality": None, "birth_place": None, "entity_type": None},
        "results": [
            {
                "source": "pep",
                "score": 100,
                "entity": {"name": "JORGE FERNÁNDEZ", "schema": "Person", "birth_dates": ["1965-03-01"], "birth_places": [], "citizenships": ["ar"], "political": ["PRIMERO SAN LUIS"], "topics": ["role.pep"]},
                "pep": {"id": "NK-1", "url": "https://opensanctions.org/entities/NK-1", "datasets": [{"id": "ar_parliament", "title": "Argentina Members of Parliament", "country": "ar", "url": "https://www.opensanctions.org/datasets/ar_parliament/"}], "matched_name": "JORGE FERNÁNDEZ", "details": [{"feature": "naam", "score": 100, "label": "Naam 100% (via \"JORGE FERNÁNDEZ\")"}]},
                "eu": None, "opensanctions": None,
            }
        ],
        "warnings": ["OpenSanctions tijdelijk niet beschikbaar"],
        "meta": {"generation_date": "2026-07-28T11:43:32", "last_modified": "Tue, 28 Jul 2026 11:00:00 GMT"},
        "pep_meta": {"updated_at": "2026-07-31T14:20:01"},
        "version": "v1.5.0",
        "author": "Dennis",
        "generated_at": "2026-07-31 15:40 CET",
        "threshold": 90,
        "max_results": 20,
    }
    payload.update(over)
    return payload


def test_escape():
    assert _escape("JORGE <FERNÁNDEZ> & \"Co\"") == "JORGE &lt;FERNÁNDEZ&gt; &amp; &quot;Co&quot;"


def _decoded_text(data: bytes) -> bytes:
    text = b""
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", data, re.S):
        stream = match.group(1)
        try:
            stream = asciiBase85Decode(stream)
        except Exception:
            pass
        try:
            stream = zlib.decompress(stream)
        except Exception:
            pass
        text += stream
    return text


def test_render_returns_pdf_bytes():
    data = render_search_pdf(_payload())
    assert data[:4] == b"%PDF"
    assert b"%%EOF" in data[-20:]


def test_render_contains_required_sections():
    data = render_search_pdf(_payload())
    # Reportlab codeert content-streams (ASCII85 + zlib); decoderen maakt tekst doorzoekbaar.
    decoded = _decoded_text(data)
    for needle in [b"JORGE FERNANDEZ", b"Uitgevoerd", b"Dennis", b"OpenSanctions", b"Disclaimer", b"90", b"EU-lijst laatste wijziging"]:
        assert needle in decoded


def test_render_empty_results():
    data = render_search_pdf(_payload(results=[], warnings=[]))
    assert data[:4] == b"%PDF"
    assert b"Geen overeenkomsten" in _decoded_text(data)


def test_render_opensanctions_result():
    payload = _payload(results=[{
        "source": "opensanctions",
        "score": 80,
        "entity": {"name": "JORGE FERNÁNDEZ", "schema": "Person"},
        "eu": None,
        "pep": None,
        "opensanctions": {
            "id": "NK-9",
            "url": "https://opensanctions.org/entities/NK-9",
            "match": True,
            "explanations": {"name_match": {"score": 0.9}},
            "datasets": ["eu_fsf"],
            "properties": {"topics": ["sanction", "role.politician"]},
        },
    }])
    data = render_search_pdf(payload)
    decoded = _decoded_text(data)
    assert data[:4] == b"%PDF"
    assert b"Match-status: match" in decoded
    assert b"Risico-tags: sanction" in decoded
    assert b"Risico-tags: role.politician" in decoded
    assert b"score 90" in decoded


def test_render_pep_result_fields():
    payload = _payload()
    decoded = _decoded_text(render_search_pdf(payload))
    assert b"PRIMERO SAN LUIS" in decoded
    assert b"role.pep" in decoded


def test_render_pep_result_positions():
    payload = _payload()
    payload["results"][0]["entity"]["positions"] = [
        {"role": "Minister van Buitenlandse Zaken", "status": "current", "start": "2021-01-01", "end": ""},
        {"role": "Senator", "status": "ended", "start": "2015-01-01", "end": "2019-01-01"},
    ]
    decoded = _decoded_text(render_search_pdf(payload))
    assert b"Functies: Minister van Buitenlandse Zaken \\(current, 2021-01-01-\\)" in decoded
    assert b"Functies: Senator \\(ended, 2015-01-01-2019-01-01\\)" in decoded


def test_render_pep_result_without_positions_omits_line():
    decoded = _decoded_text(render_search_pdf(_payload()))
    assert b"Functies:" not in decoded


def test_render_eu_result_fields():
    payload = _payload(results=[{
        "source": "eu",
        "score": 92,
        "entity": {
            "name": "ALIAS BV",
            "eu_reference_number": "EU.123",
            "united_nations_id": "UN-777",
            "aliases": ["Alias One", "Alias Two", "Alias Three", "Alias Four", "Alias Five", "Alias Six"],
            "citizenships": [{"description": "Russian Federation", "iso2": "RU"}],
            "birthdates": [{"date": "", "year": "1971", "place": "Kabul", "city": ""}],
            "regulations": [{"number_title": "2022/123", "programme": "XX", "publication_url": "https://eur-lex.europa.eu/x"}],
            "function": "Diplomat",
            "remarks": ["Opmerking test"],
        },
        "eu": {"matched_alias": "Alias One", "details": []},
        "opensanctions": None,
        "pep": None,
    }])
    decoded = _decoded_text(render_search_pdf(payload))
    assert b"Aliassen: Alias One, Alias Two, Alias Three, Alias Four, Alias Five" in decoded
    assert b"VN-id: UN-777" in decoded
    assert b"Nationaliteit: Russian Federation" in decoded
    assert b"Functie: Diplomat" in decoded
    assert b"Geboortedata/-plaats: 1971 Kabul" in decoded
    assert b"2022/123" in decoded
    assert b"Opmerking test" in decoded


def test_render_eu_result_with_empty_details():
    payload = _payload(results=[{
        "source": "eu",
        "score": 100,
        "entity": {"name": "Persoon", "eu_reference_number": "EU.1"},
        "eu": {"matched_alias": "Persoon", "details": []},
        "opensanctions": None,
        "pep": None,
    }])
    data = render_search_pdf(payload)
    assert data[:4] == b"%PDF"


def test_render_many_results_paginates():
    payload = _payload(results=[])
    for i in range(30):
        payload["results"].append({
            "source": "eu",
            "score": 100,
            "entity": {"name": f"Persoon {i}", "eu_reference_number": f"EU.{i}"},
            "eu": {"matched_alias": f"Persoon {i}", "details": [{"feature": "naam", "score": 100, "label": f"Naam 100% (via \"Persoon {i}\")"}]},
            "opensanctions": None, "pep": None,
        })
    data = render_search_pdf(payload)
    assert data[:4] == b"%PDF"


def test_render_empty_query_fields_shown_as_nvt():
    payload = _payload()
    text = _decoded_text(render_search_pdf(payload))
    assert b"Geboortejaar: NVT" in text
    assert b"Nationaliteit: NVT" in text
    assert b"Geboorteplaats: NVT" in text
    assert b"Type: NVT" in text


def test_render_filled_query_fields():
    payload = _payload(query={"name": "JORGE FERNANDEZ", "birth_year": 1965, "nationality": "AR", "birth_place": "Buenos Aires", "entity_type": "person"})
    text = _decoded_text(render_search_pdf(payload))
    assert b"Geboortejaar: 1965" in text
    assert b"Nationaliteit: AR" in text
    assert b"Geboorteplaats: Buenos Aires" in text
    assert b"Type: person" in text


def test_export_rows_pep():
    rows = _export_rows(_payload()["results"])
    assert rows == [[
        "JORGE FERNÁNDEZ", "100", "PEP", "Argentina Members of Parliament",
        'Naam 100% (via "JORGE FERNÁNDEZ")', "", "1965-03-01", "ar",
        "https://opensanctions.org/entities/NK-1",
    ]]


def test_export_rows_eu():
    result = {
        "source": "eu",
        "score": 92,
        "entity": {
            "name": "ALIAS BV",
            "eu_reference_number": "EU.123",
            "birthdates": [
                {"date": "", "year": "1971", "place": "Kabul", "city": ""},
                {"date": "1965-03-01", "year": "", "place": "", "city": ""},
            ],
            "citizenships": [{"description": "Russian Federation", "iso2": "RU"}],
        },
        "eu": {"matched_alias": "Alias One", "details": [{"feature": "naam", "score": 92, "label": 'Naam 92% (via "ALIAS BV")'}]},
        "opensanctions": None,
        "pep": None,
    }
    rows = _export_rows([result])
    assert rows == [["ALIAS BV", "92", "EU", "", 'Naam 92% (via "ALIAS BV")', "EU.123", "1971/1965-03-01", "Russian Federation", ""]]


def test_export_rows_opensanctions():
    result = {
        "source": "opensanctions",
        "score": 80,
        "entity": {"name": "JORGE FERNÁNDEZ", "schema": "Person", "birthdates": [], "citizenships": []},
        "eu": None,
        "pep": None,
        "opensanctions": {"url": "https://opensanctions.org/entities/NK-9", "datasets": ["eu_fsf"]},
    }
    rows = _export_rows([result])
    assert rows == [["JORGE FERNÁNDEZ", "80", "OpenSanctions", "eu_fsf", "", "", "", "", "https://opensanctions.org/entities/NK-9"]]


def test_export_rows_empty():
    assert _export_rows([]) == []


def test_render_csv_has_bom_header_and_data_row():
    text = render_search_csv(_payload()["results"], _payload()["query"])
    assert text.startswith("\ufeff")
    lines = text.lstrip("\ufeff").rstrip("\r\n").split("\r\n")
    assert len(lines) == 2
    assert lines[0] == "naam;score;bron;datasets;match-details;eu_referentie;geboortedata;nationaliteit;links"
    assert "JORGE FERNÁNDEZ" in lines[1]
    assert "1965-03-01" in lines[1]


def test_render_csv_special_chars_roundtrip():
    results = _payload()["results"]
    results[0]["pep"]["details"] = [
        {"feature": "naam", "score": 100, "label": 'Naam 100% (via "JORGE FERNÁNDEZ")'},
        {"feature": "land", "score": 90, "label": "Land 90%; via AR"},
    ]
    text = render_search_csv(results, _payload()["query"])
    assert "\ufeff" in text
    assert "\r\n" in text
    parsed = list(csv.reader(StringIO(text.lstrip("\ufeff")), delimiter=";"))
    assert parsed[0] == ["naam", "score", "bron", "datasets", "match-details", "eu_referentie", "geboortedata", "nationaliteit", "links"]
    assert parsed[1][0] == "JORGE FERNÁNDEZ"
    assert parsed[1][4] == 'Naam 100% (via "JORGE FERNÁNDEZ");Land 90%; via AR'


def test_render_csv_empty_results_header_only():
    text = render_search_csv([], _payload()["query"])
    assert text.startswith("\ufeff")
    lines = text.lstrip("\ufeff").rstrip("\r\n").split("\r\n")
    assert lines == ["naam;score;bron;datasets;match-details;eu_referentie;geboortedata;nationaliteit;links"]


def test_render_xlsx_magic_cells_and_widths():
    data = render_search_xlsx(_payload()["results"], _payload()["query"])
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["naam", "score", "bron", "datasets", "match-details", "eu_referentie", "geboortedata", "nationaliteit", "links"]
    assert ws.cell(row=1, column=1).font.bold
    row = [c.value for c in ws[2]]
    assert row[0] == "JORGE FERNÁNDEZ"
    assert row[1] == "100"
    assert row[2] == "PEP"
    assert row[8] == "https://opensanctions.org/entities/NK-1"
    assert ws.column_dimensions["A"].width > 10


def test_render_xlsx_empty_results_header_only():
    data = render_search_xlsx([], _payload()["query"])
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["naam", "score", "bron", "datasets", "match-details", "eu_referentie", "geboortedata", "nationaliteit", "links"]
    assert ws.max_row == 1
