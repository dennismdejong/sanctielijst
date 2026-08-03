import pytest
from fastapi.testclient import TestClient

from app import search_index
from app.main import create_app


@pytest.fixture(autouse=True)
def _isolate_env(monkeypatch, tmp_path):
    monkeypatch.delenv("OPENSANCTIONS_API_KEY", raising=False)
    monkeypatch.setenv("SEARCH_DB", str(tmp_path / "no-search.sqlite"))
    monkeypatch.setenv("EU_DATA_DIR", str(tmp_path / "eu"))
    monkeypatch.setenv(search_index.INDEX_ENV, "0")
    monkeypatch.setenv("AUDIT_DB", str(tmp_path / "audit.sqlite"))
    monkeypatch.setenv("BATCH_DB", str(tmp_path / "batch.sqlite"))
    monkeypatch.setenv("WATCHLIST_DB", str(tmp_path / "watchlists.sqlite"))
    monkeypatch.delenv("AUDIT_ADMIN_TOKEN", raising=False)


def make_entity(eu_ref, whole_name, subject_type="person", year=None, country=None, place=None):
    return {
        "logical_id": eu_ref,
        "eu_reference_number": eu_ref,
        "united_nations_id": "",
        "designation_date": "2022-01-01",
        "subject_type": subject_type,
        "aliases": [{"whole_name": whole_name, "first_name": "", "last_name": "", "strong": True, "function": "Diplomat", "title": ""}],
        "citizenships": [{"iso2": country, "description": country}] if country else [],
        "birthdates": [{"date": "", "year": year, "year_from": None, "year_to": None, "city": "", "place": place, "iso2": country or "", "country": country or ""}] if year or place else [],
        "addresses": [],
        "identifications": [],
        "regulations": [{"number_title": "2022/123", "publication_date": "2022-02-01", "programme": "XX", "publication_url": "https://eur-lex.europa.eu/x"}],
        "remarks": [],
    }


ENTITIES = [
    make_entity("EU.471.56", "Abdul Hai Hazem Abdul Qader", year=1971, country="AF", place="Kabul"),
    make_entity("EU.2", "Rosneft", subject_type="enterprise", country="RU"),
]

def test_health():
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/health").json() == {"status": "ok"}


def test_status_fields():
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/status").json()
    assert data["entity_count"] == 2
    assert data["opensanctions_active"] is False
    assert "source" in data
    assert data["index"]["status"] == "disabled"
    assert data["index"]["enabled"] is False


def test_search_returns_eu_result():
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/search", params={"name": "Abdul Hai Hazem"}).json()
    assert data["results"]
    first = data["results"][0]
    assert first["source"] == "eu"
    assert first["eu"]["total_score"] == 100
    assert first["entity"]["eu_reference_number"] == "EU.471.56"
    assert any(d["feature"] == "naam" for d in first["eu"]["details"])


def test_search_birth_year_boosts():
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/search", params={"name": "Abdul", "birth_year": 1971}).json()
    scores = [r["score"] for r in data["results"] if r["source"] == "eu"]
    assert scores and scores[0] == 100


def test_search_entity_type_filter():
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/search", params={"name": "Rosneft", "entity_type": "enterprise"}).json()
    assert any(r["source"] == "eu" for r in data["results"])


def test_search_requires_name():
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/search")
    assert resp.status_code == 422


def test_search_whitespace_name_returns_422():
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/search", params={"name": "   "})
    assert resp.status_code == 422


def test_os_api_key_defaults_from_env(monkeypatch):
    monkeypatch.setenv("OPENSANCTIONS_API_KEY", "KEY")
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/status").json()
    assert data["opensanctions_active"] is True


def test_search_with_opensanctions_merges(monkeypatch):
    import app.main as main

    fake_os = [
        {
            "id": "NK-x",
            "caption": "Abdul Qader",
            "schema": "Person",
            "score": 0.8,
            "match": True,
            "explanations": {"name_match": {"score": 0.8}},
            "datasets": ["eu_fsf"],
            "properties": {},
            "url": "https://opensanctions.org/entities/NK-x",
        }
    ]
    monkeypatch.setattr(main.opensanctions, "match_opensanctions", lambda *a, **k: fake_os)
    client = TestClient(create_app(entities=ENTITIES, os_api_key="KEY"))
    data = client.get("/api/search", params={"name": "Abdul Qader"}).json()
    sources = {r["source"] for r in data["results"]}
    assert sources == {"eu", "opensanctions"}
    os_result = [r for r in data["results"] if r["source"] == "opensanctions"][0]
    assert os_result["score"] == 80
    assert os_result["opensanctions"]["url"] == "https://opensanctions.org/entities/NK-x"


def test_search_opensanctions_failure_adds_warning(monkeypatch):
    import app.main as main

    monkeypatch.setattr(main.opensanctions, "match_opensanctions", lambda *a, **k: (_ for _ in ()).throw(RuntimeError("down")))
    client = TestClient(create_app(entities=ENTITIES, os_api_key="KEY"))
    data = client.get("/api/search", params={"name": "Abdul Hai Hazem"}).json()
    assert data["results"]
    assert any("OpenSanctions" in w for w in data["warnings"])


def test_search_no_match_returns_empty():
    client = TestClient(create_app(entities=ENTITIES))
    data = client.get("/api/search", params={"name": "Zzq Qqxx"}).json()
    assert data["results"] == []


def test_index_serves_html(tmp_path):
    static = tmp_path / "static"
    static.mkdir()
    (static / "index.html").write_text("<h1>hi</h1>")
    client = TestClient(create_app(entities=ENTITIES, static_dir=static))
    resp = client.get("/")
    assert resp.status_code == 200
    assert resp.text == "<h1>hi</h1>"


def test_audit_page_served(tmp_path):
    static = tmp_path / "static"
    static.mkdir()
    (static / "audit.html").write_text("<h1>audit</h1>")
    client = TestClient(create_app(entities=ENTITIES, static_dir=static))
    resp = client.get("/audit")
    assert resp.status_code == 200
    assert resp.text == "<h1>audit</h1>"


def _write_pep_fixture(root):
    import json
    for ds, entities in [
        ("ar_parliament", [
            {"id": "NK-x", "caption": "JORGE FERNÁNDEZ", "schema": "Person", "target": True, "datasets": ["ar_parliament"],
             "properties": {"birthDate": ["1965-03-01"], "citizenship": ["ar"], "political": ["PRIMERO SAN LUIS"], "topics": ["role.pep"]}},
        ]),
    ]:
        p = root / ds / "entities.ftm.json"
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w") as fh:
            for e in entities:
                fh.write(json.dumps(e) + "\n")
    (root / "datasets.json").write_text(json.dumps({"ar_parliament": {"title": "Argentina Members of Parliament", "publisher": "HCDN", "country": "ar", "official": True, "url": "https://parlament.ar"}}))


def _write_pep_fixture_with_positions(root):
    import json
    for ds, entities in [
        ("ar_parliament", [
            {"id": "NK-x", "caption": "JORGE FERNÁNDEZ", "schema": "Person", "target": True, "datasets": ["ar_parliament"],
             "properties": {"birthDate": ["1965-03-01"], "citizenship": ["ar"], "political": ["PRIMERO SAN LUIS"], "topics": ["role.pep"]}},
            {"id": "P-1", "caption": "Minister of Defence", "schema": "Position", "target": False, "datasets": ["ar_parliament"],
             "properties": {"name": ["Minister of Defence"]}},
            {"id": "O-1", "caption": "Occupancy", "schema": "Occupancy", "target": False, "datasets": ["ar_parliament"],
             "properties": {"holder": ["NK-x"], "post": ["P-1"], "status": ["current"], "startDate": ["2020-01-01"], "endDate": ["2024-01-01"]}},
        ]),
    ]:
        p = root / ds / "entities.ftm.json"
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w") as fh:
            for e in entities:
                fh.write(json.dumps(e) + "\n")
    (root / "datasets.json").write_text(json.dumps({"ar_parliament": {"title": "Argentina Members of Parliament", "publisher": "HCDN", "country": "ar", "official": True, "url": "https://parlament.ar"}}))


def make_eu_entity():
    return {
        "logical_id": "EU.1", "eu_reference_number": "EU.1", "united_nations_id": "",
        "designation_date": "2022-01-01", "subject_type": "person",
        "aliases": [{"whole_name": "John Smith", "first_name": "", "last_name": "", "strong": True, "function": "", "title": ""}],
        "citizenships": [], "birthdates": [], "addresses": [], "identifications": [],
        "regulations": [{"number_title": "2022/123", "publication_date": "2022-02-01", "programme": "XX", "publication_url": "https://eur-lex.europa.eu/x"}],
        "remarks": [],
    }


def _write_search_db(root):
    _write_pep_fixture(root)
    return search_index.build_index(root / "search.sqlite", [make_eu_entity()], root)


def build_index(db_path, eu_entities, root):
    return search_index.build_index(db_path, eu_entities, root)


def _decoded_text(data: bytes) -> bytes:
    import re
    import zlib

    from reportlab.pdfbase.pdfutils import asciiBase85Decode

    text = b""
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", data, re.S):
        stream = match.group(1)
        try:
            stream = asciiBase85Decode(stream)
        except Exception:
            pass
        try:
            stream = zlib.decompress(stream)
        except Exception:
            pass
        text += stream
    return text


def test_status_index_ready(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    data = client.get("/api/status").json()
    assert data["index"]["status"] == "ready"
    assert data["index"]["pep_count"] == 1
    assert data["index"]["eu_count"] == 1


def test_search_while_building_serves_eu(tmp_path, monkeypatch):
    import threading

    import app.main as main

    eu_root = tmp_path / "eu"
    eu_root.mkdir()
    (eu_root / "eu_sanctions.xml").write_bytes(b"<export/>")
    started = threading.Event()
    release = threading.Event()

    def slow_rebuild(db_path, eu_xml, pep_root, sanctions_root):
        started.set()
        release.wait(5)

    monkeypatch.setattr(main.search_index, "rebuild_index", slow_rebuild)
    client = TestClient(create_app(entities=ENTITIES, eu_root=eu_root, search_db=tmp_path / "missing.sqlite"))
    started.wait(5)
    status = client.get("/api/status").json()
    assert status["index"]["status"] == "building"
    data = client.get("/api/search", params={"name": "Abdul Hai Hazem"}).json()
    assert data["results"]
    assert data["results"][0]["source"] == "eu"
    assert data["results"][0]["entity"]["eu_reference_number"] == "EU.471.56"
    assert any("opgebouwd" in w for w in data["warnings"])
    release.set()


def test_create_app_corrupt_db_does_not_raise(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    (tmp_path / "eu_sanctions.xml").write_bytes(b"<export/>")
    (tmp_path / "search.sqlite").write_bytes(b"kapot")
    app = create_app(entities=ENTITIES, eu_root=tmp_path, search_db=tmp_path / "search.sqlite")
    client = TestClient(app)
    data = client.get("/api/status").json()
    assert data["index"]["status"] in ("building", "ready")


def test_search_db_merges_eu_and_pep(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    data = client.get("/api/search", params={"name": "JORGE FERNANDEZ"}).json()
    assert [r["source"] for r in data["results"]] == ["pep"]
    first = [r for r in data["results"] if r["source"] == "pep"][0]
    assert first["pep"]["datasets"][0]["id"] == "ar_parliament"
    data = client.get("/api/search", params={"name": "John Smith"}).json()
    eu = [r for r in data["results"] if r["source"] == "eu"][0]
    assert eu["eu"]["matched_alias"] == "John Smith"


def test_search_pep_result_contains_positions(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture_with_positions(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    data = client.get("/api/search", params={"name": "JORGE FERNANDEZ"}).json()
    pep = [r for r in data["results"] if r["source"] == "pep"][0]
    assert pep["entity"]["positions"] == [
        {"role": "Minister of Defence", "status": "current", "start": "2020-01-01", "end": "2024-01-01"}
    ]


def test_search_pep_result_positions_default_empty(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    data = client.get("/api/search", params={"name": "JORGE FERNANDEZ"}).json()
    pep = [r for r in data["results"] if r["source"] == "pep"][0]
    assert pep["entity"]["positions"] == []


def test_search_eu_result_has_no_positions(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    data = client.get("/api/search", params={"name": "John Smith"}).json()
    eu = [r for r in data["results"] if r["source"] == "eu"][0]
    assert "positions" not in eu["entity"]


def test_default_pep_root_uses_env(monkeypatch):
    from pathlib import Path

    from app import main as main_module
    monkeypatch.setenv("PEP_DATA_DIR", "/data/pep")
    assert main_module.default_pep_root() == Path("/data/pep")
    monkeypatch.delenv("PEP_DATA_DIR", raising=False)
    assert main_module.default_pep_root() == main_module.PEP_ROOT


def test_default_eu_root_uses_env(monkeypatch):
    from pathlib import Path

    from app import main as main_module
    monkeypatch.setenv("EU_DATA_DIR", "/data/eu")
    assert main_module.default_eu_root() == Path("/data/eu")


def test_default_eu_root_falls_back_without_env(monkeypatch):
    from app import main as main_module
    monkeypatch.delenv("EU_DATA_DIR", raising=False)
    assert main_module.default_eu_root() == main_module.EU_ROOT


def test_refresh_success(tmp_path, monkeypatch):
    import app.main as main

    eu_root = tmp_path / "eu"
    eu_root.mkdir()
    manifest = {"status": "ok", "downloaded_at": "2026-07-31T12:00:00+00:00", "generation_date": "2026-07-28T11:43:32+02:00", "entity_count": 2}
    monkeypatch.setattr(main.eu_ingest, "refresh_eu", lambda *a, **k: manifest)
    client = TestClient(create_app(entities=ENTITIES, eu_root=eu_root))
    resp = client.post("/api/refresh")
    assert resp.status_code == 200
    assert resp.json()["source"] == "ok"


def test_refresh_head_failure_returns_503(tmp_path, monkeypatch):
    import app.main as main

    eu_root = tmp_path / "eu"
    eu_root.mkdir()
    monkeypatch.setattr(main.eu_ingest, "refresh_eu", lambda *a, **k: (_ for _ in ()).throw(RuntimeError("down")))
    client = TestClient(create_app(entities=ENTITIES, eu_root=eu_root))
    resp = client.post("/api/refresh")
    assert resp.status_code == 503


def test_status_source_from_manifest(tmp_path):
    import json

    eu_root = tmp_path / "eu"
    eu_root.mkdir()
    (eu_root / "manifest.json").write_text(json.dumps({
        "status": "ok",
        "downloaded_at": "2026-07-31T12:00:00+00:00",
        "generation_date": "2026-07-28T11:43:32+02:00",
    }))
    client = TestClient(create_app(entities=ENTITIES, eu_root=eu_root))
    data = client.get("/api/status").json()
    assert data["source"] == "ok"
    assert data["generated_at"] == "2026-07-28T11:43:32+02:00"
    assert data["entity_count"] == 2


def test_startup_corrupt_xml_boots_with_error(tmp_path):
    import json

    (tmp_path / "eu_sanctions.xml").write_bytes(b"<garbage")
    (tmp_path / "manifest.json").write_text(json.dumps({"status": "ok"}))
    client = TestClient(create_app(eu_root=tmp_path))
    resp = client.get("/api/status")
    assert resp.status_code == 200
    data = resp.json()
    assert data["source"] == "error"
    assert data["entity_count"] == 0


def test_status_data_age_hours_tz_naive_ok(tmp_path):
    import json

    from app import main as main_module

    eu_root = tmp_path / "eu"
    eu_root.mkdir()
    (eu_root / "manifest.json").write_text(json.dumps({
        "status": "ok",
        "downloaded_at": "2026-07-31T12:00:00",
    }))
    client = TestClient(create_app(entities=ENTITIES, eu_root=eu_root))
    data = client.get("/api/status").json()
    assert data["source"] == "ok"
    assert data["data_age_hours"] is not None
    assert isinstance(data["data_age_hours"], float)
    assert main_module._data_age_hours("garbage") is None
    assert main_module._data_age_hours(None) is None


def test_export_returns_pdf(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "JORGE FERNANDEZ", "author": "Dennis"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:4] == b"%PDF"


def test_export_csv_format(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "JORGE FERNANDEZ", "format": "csv"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:3] == b"\xef\xbb\xbf"
    assert resp.content[3:6] != b"\xef\xbb\xbf"
    body = resp.content.decode("utf-8-sig")
    assert not body.startswith("\ufeff")
    assert "naam;score;bron;datasets;match-details;eu_referentie;geboortedata;nationaliteit;links" in body
    assert "JORGE FERNÁNDEZ" in body
    assert b".csv" in resp.headers["content-disposition"].encode()


def test_export_xlsx_format(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "JORGE FERNANDEZ", "format": "xlsx"})
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(resp.content))
    ws = wb["Screening"]
    assert [c.value for c in ws[1]] == ["naam", "score", "bron", "datasets", "match-details", "eu_referentie", "geboortedata", "nationaliteit", "links"]
    assert ws["A2"].value == "JORGE FERNÁNDEZ"
    assert b".xlsx" in resp.headers["content-disposition"].encode()


def test_export_invalid_format_is_422(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "JORGE FERNANDEZ", "format": "onzin"})
    assert resp.status_code == 422


def test_export_requires_name():
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/search/export").status_code == 422


def test_export_empty_results(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "Zzqqq Xxww"})
    assert resp.status_code == 200
    assert b"Geen overeenkomsten" in _decoded_text(resp.content)


def _last_audit_events(tmp_path):
    from app import audit

    return audit.list_events(tmp_path / "audit.sqlite")


def test_search_logs_audit_event(tmp_path):
    from app import matcher

    client = TestClient(create_app(entities=ENTITIES))
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    events = _last_audit_events(tmp_path)
    assert len(events) == 1
    event = events[0]
    assert event["ip"] == "testclient"
    assert event["user"] is None
    assert event["method"] == "GET"
    assert event["path"] == "/api/search"
    assert event["query"] == {"name": "Abdul Hai Hazem", "birth_year": None, "nationality": None, "birth_place": None, "entity_type": None}
    assert event["result_count"] >= 1
    assert event["sources"] == ["eu"]
    assert event["threshold"] == matcher.THRESHOLD
    assert event["user_agent"]


def test_search_logs_empty_results(tmp_path):
    client = TestClient(create_app(entities=ENTITIES))
    client.get("/api/search", params={"name": "Zzq Qqxx"})
    event = _last_audit_events(tmp_path)[0]
    assert event["result_count"] == 0
    assert event["sources"] == []


def test_export_logs_audit_event(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    build_index(tmp_path / "search.sqlite", [make_eu_entity()], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.get("/api/search/export", params={"name": "JORGE FERNANDEZ"})
    assert resp.status_code == 200
    event = _last_audit_events(tmp_path)[0]
    assert event["path"] == "/api/search/export"
    assert event["result_count"] >= 1


def test_audit_endpoint_disabled_without_token():
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/audit")
    assert resp.status_code == 404


def test_audit_logs_422_blank_name(monkeypatch):
    monkeypatch.setenv("AUDIT_ADMIN_TOKEN", "secret")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/search", params={"name": "   "}).status_code == 422
    assert client.get("/api/search/export", params={"name": "   "}).status_code == 422
    resp = client.get("/api/audit", headers={"Authorization": "Bearer secret"})
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert all(e["result_count"] == 0 for e in data["events"])


def test_audit_endpoint_401_without_or_bad_token(monkeypatch):
    monkeypatch.setenv("AUDIT_ADMIN_TOKEN", "secret")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/audit").status_code == 401
    assert client.get("/api/audit", headers={"Authorization": "Bearer wrong"}).status_code == 401


def test_audit_endpoint_returns_events_with_token(monkeypatch):
    monkeypatch.setenv("AUDIT_ADMIN_TOKEN", "secret")
    client = TestClient(create_app(entities=ENTITIES))
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    resp = client.get("/api/audit", headers={"Authorization": "Bearer secret"})
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 1
    assert len(data["events"]) == 1
    assert data["events"][0]["ip"] == "testclient"
    assert data["events"][0]["query"]["name"] == "Abdul Hai Hazem"


def test_audit_endpoint_pagination(monkeypatch):
    monkeypatch.setenv("AUDIT_ADMIN_TOKEN", "secret")
    client = TestClient(create_app(entities=ENTITIES))
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    client.get("/api/search", params={"name": "Rosneft"})
    resp = client.get("/api/audit", params={"limit": 1, "offset": 1}, headers={"Authorization": "Bearer secret"})
    data = resp.json()
    assert data["total"] == 2
    assert len(data["events"]) == 1


def test_audit_failure_does_not_break_search(monkeypatch):
    import app.audit as audit_module

    def boom(*args, **kwargs):
        raise RuntimeError("disk vol")

    monkeypatch.setattr(audit_module, "log_event", boom)
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    assert resp.status_code == 200
    assert resp.json()["results"]


@pytest.fixture
def auth_env(monkeypatch, tmp_path):
    monkeypatch.setenv("AUTH_SECRET", "test-secret")
    monkeypatch.setenv("AUTH_DB", str(tmp_path / "auth.sqlite"))


def _create_local_user(username="alice", password="geheim", role="admin"):
    from app import auth

    return auth.create_user(auth.default_auth_db(), username=username, password=password, role=role)


def _set_entra_env(monkeypatch, **overrides):
    env = {
        "AUTH_ENTRA_ENABLED": "1",
        "AUTH_ENTRA_TENANT": "tenant-test",
        "AUTH_ENTRA_CLIENT_ID": "client-1",
        "AUTH_ENTRA_CLIENT_SECRET": "secret-1",
        "AUTH_ENTRA_REDIRECT_URI": "https://app.example/callback",
    }
    env.update(overrides)
    for key, value in env.items():
        monkeypatch.setenv(key, value)


def test_auth_login_get_returns_local_methods():
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/auth/login")
    assert resp.status_code == 200
    assert resp.json() == {"methods": ["local"]}


def test_auth_login_local_sets_cookie_and_me(auth_env):
    _create_local_user()
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    assert resp.status_code == 200
    assert resp.json() == {"username": "alice", "role": "admin"}
    set_cookie = resp.headers.get("set-cookie", "")
    assert "session=" in set_cookie
    assert "HttpOnly" in set_cookie
    assert "SameSite=lax" in set_cookie
    assert "Path=/" in set_cookie
    assert "Secure" not in set_cookie
    me = client.get("/api/auth/me")
    assert me.status_code == 200
    assert me.json() == {"username": "alice", "role": "admin"}


def test_auth_login_local_bad_credentials_401(auth_env):
    _create_local_user()
    client = TestClient(create_app(entities=ENTITIES))
    assert client.post("/api/auth/login", json={"username": "alice", "password": "fout"}).status_code == 401
    assert client.post("/api/auth/login", json={"username": "onbekend", "password": "geheim"}).status_code == 401
    assert client.post("/api/auth/login", json={}).status_code == 401


def test_auth_logout_clears_cookie(auth_env):
    _create_local_user()
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    assert client.get("/api/auth/me").status_code == 200
    resp = client.post("/api/auth/logout")
    assert resp.status_code == 204
    assert client.get("/api/auth/me").status_code == 401


def test_auth_me_401_when_anonymous():
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/auth/me").status_code == 401


def test_auth_login_entra_redirects(monkeypatch, auth_env):
    import app.main as main

    _set_entra_env(monkeypatch)
    fake_url = "https://login.microsoftonline.com/tenant-test/oauth2/v2.0/authorize?x=1"

    async def fake_authorize(*a, **k):
        return (fake_url, "verifier", "state-1")

    monkeypatch.setattr(main.auth, "entra_authorize_url", fake_authorize)
    client = TestClient(create_app(entities=ENTITIES), follow_redirects=False)
    resp = client.get("/api/auth/login")
    assert resp.status_code == 307
    assert resp.headers["location"] == fake_url
    assert "auth_code_verifier=verifier" in resp.headers.get("set-cookie", "")


def test_auth_callback_entra_success(monkeypatch, auth_env):
    import app.main as main

    _set_entra_env(monkeypatch)
    userinfo = {"sub": "sub-123", "username": "bob@example.com", "preferred_username": "bob@example.com", "email": "bob@example.com"}

    async def fake_exchange(client, code, code_verifier, state, state_secret):
        return userinfo

    monkeypatch.setattr(main.auth, "entra_exchange", fake_exchange)
    client = TestClient(create_app(entities=ENTITIES), follow_redirects=False)
    client.cookies.set("auth_code_verifier", "verifier")
    resp = client.get("/api/auth/callback", params={"code": "code-1", "state": "state-1"})
    assert resp.status_code == 303
    assert resp.headers["location"] == "/"
    me = client.get("/api/auth/me")
    assert me.status_code == 200
    assert me.json() == {"username": "bob@example.com", "role": "viewer"}


def test_auth_callback_invalid_state_400(monkeypatch, auth_env):
    import app.main as main

    _set_entra_env(monkeypatch)

    async def fake_exchange(client, code, code_verifier, state, state_secret):
        raise ValueError("ongeldige of verlopen state")

    monkeypatch.setattr(main.auth, "entra_exchange", fake_exchange)
    client = TestClient(create_app(entities=ENTITIES))
    client.cookies.set("auth_code_verifier", "verifier")
    resp = client.get("/api/auth/callback", params={"code": "code-1", "state": "onzin"})
    assert resp.status_code == 400


def test_auth_callback_local_username_collision_400(monkeypatch, auth_env):
    import app.main as main

    _set_entra_env(monkeypatch)
    _create_local_user(username="bob@example.com", role="viewer")
    userinfo = {"sub": "sub-bob", "username": "bob@example.com", "preferred_username": "bob@example.com", "email": "bob@example.com"}

    async def fake_exchange(client, code, code_verifier, state, state_secret):
        return userinfo

    monkeypatch.setattr(main.auth, "entra_exchange", fake_exchange)
    client = TestClient(create_app(entities=ENTITIES))
    client.cookies.set("auth_code_verifier", "verifier")
    resp = client.get("/api/auth/callback", params={"code": "code-1", "state": "state-1"})
    assert resp.status_code == 400
    assert resp.json()["detail"] == "Gebruiker bestaat al"


def test_auth_callback_missing_code_verifier_400(monkeypatch, auth_env):
    import app.main as main

    _set_entra_env(monkeypatch)
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/auth/callback", params={"code": "code-1", "state": "state-1"})
    assert resp.status_code == 400


def test_auth_callback_entra_disabled_400():
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.get("/api/auth/callback", params={"code": "code-1", "state": "state-1"})
    assert resp.status_code == 400


def test_startup_raises_when_required_without_secret(monkeypatch):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    monkeypatch.delenv("AUTH_SECRET", raising=False)
    with pytest.raises(RuntimeError):
        create_app(entities=ENTITIES)


def test_startup_raises_when_entra_config_incomplete(monkeypatch, auth_env):
    _set_entra_env(monkeypatch, AUTH_ENTRA_CLIENT_SECRET="")
    with pytest.raises(RuntimeError):
        create_app(entities=ENTITIES)


def test_startup_raises_when_entra_enabled_without_secret(monkeypatch):
    _set_entra_env(monkeypatch)
    monkeypatch.delenv("AUTH_SECRET", raising=False)
    with pytest.raises(RuntimeError):
        create_app(entities=ENTITIES)


def test_gating_requires_login_when_required(monkeypatch, auth_env):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/search", params={"name": "Abdul"}).status_code == 401
    assert client.get("/api/search/export", params={"name": "Abdul"}).status_code == 401


def test_gating_allows_authenticated_user(monkeypatch, auth_env):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    _create_local_user(role="analist")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.post("/api/auth/login", json={"username": "alice", "password": "geheim"}).status_code == 200
    assert client.get("/api/search", params={"name": "Abdul Hai Hazem"}).status_code == 200
    assert client.get("/api/search/export", params={"name": "Zzq Qqxx"}).status_code == 200


def test_gating_viewer_cannot_export(monkeypatch, auth_env):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    _create_local_user(role="viewer")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.post("/api/auth/login", json={"username": "alice", "password": "geheim"}).status_code == 200
    assert client.get("/api/search", params={"name": "Abdul Hai Hazem"}).status_code == 200
    assert client.get("/api/search/export", params={"name": "Abdul Hai Hazem"}).status_code == 403
    assert client.get("/api/search/export", params={"name": "Abdul Hai Hazem", "format": "csv"}).status_code == 403


def test_gating_open_when_not_required(auth_env):
    _create_local_user()
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/search", params={"name": "Abdul Hai Hazem"}).status_code == 200


def test_users_endpoint_requires_admin(auth_env):
    _create_local_user(username="admin", role="admin")
    _create_local_user(username="analist", role="analist")
    client = TestClient(create_app(entities=ENTITIES))
    payload = {"username": "newbie", "password": "x", "role": "viewer"}
    assert client.post("/api/auth/users", json=payload).status_code == 401
    client.post("/api/auth/login", json={"username": "analist", "password": "geheim"})
    assert client.post("/api/auth/users", json=payload).status_code == 403
    client.post("/api/auth/login", json={"username": "admin", "password": "geheim"})
    resp = client.post("/api/auth/users", json=payload)
    assert resp.status_code == 200
    assert resp.json()["username"] == "newbie"
    assert resp.json()["role"] == "viewer"


def test_users_endpoint_creates_entra_user(auth_env):
    _create_local_user(role="admin")
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    resp = client.post("/api/auth/users", json={"username": "bob@example.com", "role": "analist", "idp_subject": "sub-1"})
    assert resp.status_code == 200
    assert resp.json()["username"] == "bob@example.com"
    assert resp.json()["role"] == "analist"


def test_users_endpoint_duplicate_400(auth_env):
    _create_local_user(role="admin")
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    payload = {"username": "newbie", "password": "x", "role": "viewer"}
    assert client.post("/api/auth/users", json=payload).status_code == 200
    assert client.post("/api/auth/users", json=payload).status_code == 400


def test_audit_event_user_filled_when_logged_in(auth_env, tmp_path):
    _create_local_user(role="analist")
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    event = _last_audit_events(tmp_path)[0]
    assert event["user"] == "alice"


def test_audit_event_user_none_when_anonymous(tmp_path):
    client = TestClient(create_app(entities=ENTITIES))
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    event = _last_audit_events(tmp_path)[0]
    assert event["user"] is None


def test_audit_endpoint_allows_admin_role(auth_env):
    _create_local_user(role="admin")
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    client.get("/api/search", params={"name": "Abdul Hai Hazem"})
    resp = client.get("/api/audit")
    assert resp.status_code == 200
    assert resp.json()["total"] == 1


def test_build_index_runs_rebuild_in_subprocess(tmp_path, monkeypatch):
    import subprocess as sp
    import sys as _sys

    import app.main as main

    captured = {}

    def fake_run(cmd, **kwargs):
        captured["cmd"] = cmd
        return sp.CompletedProcess(cmd, 0, stdout='{"eu_count": 1, "pep_count": 2, "total": 3, "source_count": 1}', stderr="")

    monkeypatch.setenv("PEP_INDEX_SUBPROCESS", "1")
    monkeypatch.setattr(main.subprocess, "run", fake_run)
    state = {"index_status": "building", "index_stats": None, "index_error": None}
    main._build_index(state, tmp_path / "db.sqlite", tmp_path / "eu.xml", tmp_path, tmp_path / "sanc")
    assert state["index_status"] == "ready"
    assert state["index_stats"]["total"] == 3
    assert captured["cmd"][0] == _sys.executable
    assert "app.rebuild" in captured["cmd"]


def test_build_index_subprocess_failure_sets_error(tmp_path, monkeypatch):
    import subprocess as sp

    import app.main as main

    def fake_run(cmd, **kwargs):
        return sp.CompletedProcess(cmd, 1, stdout="", stderr="boom")

    monkeypatch.setenv("PEP_INDEX_SUBPROCESS", "1")
    monkeypatch.setattr(main.subprocess, "run", fake_run)
    state = {"index_status": "building", "index_stats": None, "index_error": None}
    main._build_index(state, tmp_path / "db.sqlite", tmp_path / "eu.xml", tmp_path, tmp_path / "sanc")
    assert state["index_status"] == "error"
    assert "boom" in state["index_error"]


def test_status_triggers_rebuild_when_data_newer(tmp_path, monkeypatch):
    import os
    import threading
    import time

    import app.main as main

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    called = threading.Event()
    release = threading.Event()

    def counting_rebuild(db_path, eu_xml, pep_root, sanctions_root):
        called.set()
        release.wait(5)

    monkeypatch.setattr(main.search_index, "rebuild_index", counting_rebuild)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    assert client.get("/api/status").json()["index"]["status"] == "ready"
    future = time.time() + 1000
    os.utime(tmp_path / "ar_parliament" / "entities.ftm.json", (future, future))
    data = client.get("/api/status").json()
    assert data["index"]["status"] == "building"
    assert called.wait(5)
    release.set()


def test_status_no_endless_rebuild_with_future_input_mtime(tmp_path, monkeypatch):
    import os
    import threading
    import time

    import app.main as main

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    builds = []
    done = threading.Event()
    original_rebuild = search_index.rebuild_index

    def real_rebuild(db_path, eu_xml, pep_root, sanctions_root):
        builds.append(1)
        stats = original_rebuild(db_path, eu_xml, pep_root)
        done.set()
        return stats

    monkeypatch.setattr(main.search_index, "rebuild_index", real_rebuild)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    assert client.get("/api/status").json()["index"]["status"] == "ready"
    future = time.time() + 1000
    os.utime(tmp_path / "ar_parliament" / "entities.ftm.json", (future, future))
    assert client.get("/api/status").json()["index"]["status"] == "building"
    assert done.wait(10)
    deadline = time.time() + 10
    while time.time() < deadline:
        if client.get("/api/status").json()["index"]["status"] == "ready":
            break
        time.sleep(0.02)
    for _ in range(3):
        assert client.get("/api/status").json()["index"]["status"] == "ready"
    assert builds == [1]


def test_build_index_subprocess_timeout_sets_error(tmp_path, monkeypatch):
    import subprocess as sp

    import app.main as main

    def fake_run(cmd, **kwargs):
        raise sp.TimeoutExpired(cmd=cmd, timeout=kwargs.get("timeout", 600), output="", stderr="bouw hangt")

    monkeypatch.setenv("PEP_INDEX_SUBPROCESS", "1")
    monkeypatch.setattr(main.subprocess, "run", fake_run)
    state = {"index_status": "building", "index_stats": None, "index_error": None}
    main._build_index(state, tmp_path / "db.sqlite", tmp_path / "eu.xml", tmp_path, tmp_path / "sanc")
    assert state["index_status"] == "error"
    assert "timeout" in state["index_error"].lower()
    assert "600" in state["index_error"]
    assert "bouw hangt" in state["index_error"]


def _batch_csv(text: str) -> bytes:
    return text.encode("utf-8")


def _create_batch(client, text: str) -> str:
    resp = client.post("/api/batch", files={"file": ("lijst.csv", _batch_csv(text), "text/csv")})
    assert resp.status_code == 200, resp.text
    return resp.json()["batch_id"]


def _wait_done(client, batch_id: str, timeout: float = 10.0) -> dict:
    import time

    deadline = time.time() + timeout
    while time.time() < deadline:
        data = client.get(f"/api/batch/{batch_id}").json()
        if data["status"] in ("done", "error"):
            return data
        time.sleep(0.05)
    raise AssertionError(f"batch {batch_id} niet klaar binnen {timeout}s")


def test_batch_upload_returns_batch_id(tmp_path, monkeypatch):
    import uuid

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    uuid.UUID(batch_id)


def test_batch_processes_rows_and_reports_matches(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\nJohn Smith\nZzq Qqxx\n")
    status = client.get(f"/api/batch/{batch_id}").json()
    assert status["status"] in ("pending", "running", "done")
    assert status["total"] == 3
    data = _wait_done(client, batch_id)
    assert data["status"] == "done"
    assert data["progress"] == 3
    assert data["finished_at"] is not None
    by_name = {r["row"]["naam"]: r for r in data["rows"]}
    jorge = by_name["JORGE FERNANDEZ"]
    assert jorge["matches"] and jorge["matches"][0]["source"] == "pep"
    john = by_name["John Smith"]
    assert john["matches"] and john["matches"][0]["source"] == "eu"
    assert by_name["Zzq Qqxx"]["matches"] == []


def test_batch_unknown_job_returns_404():
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/batch/onbekend").status_code == 404


def test_batch_surfaces_per_row_parse_errors(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam;geboortejaar\nJan Jansen;1970\n;1980\n")
    data = _wait_done(client, batch_id)
    assert data["status"] == "done"
    assert data["total"] == 1
    assert data["errors"] == [{"row_index": 3, "error": "Ontbrekende naam"}]
    assert [r["row"]["naam"] for r in data["rows"]] == ["Jan Jansen"]


def test_batch_row_limit_413(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    text = "naam\n" + "\n".join(f"persoon-{i}" for i in range(5001)) + "\n"
    resp = client.post("/api/batch", files={"file": ("lijst.csv", _batch_csv(text), "text/csv")})
    assert resp.status_code == 413


def test_batch_invalid_birth_year_csv_is_per_row_error(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam;geboortejaar\nJORGE FERNANDEZ;onbekend\nJohn Smith;1971\n")
    data = _wait_done(client, batch_id)
    assert data["status"] == "done"
    assert data["errors"] == [{"row_index": 2, "error": "Ongeldig geboortejaar"}]
    by_name = {r["row"]["naam"]: r for r in data["rows"]}
    assert by_name["JORGE FERNANDEZ"]["row"]["geboortejaar"] is None
    assert by_name["JORGE FERNANDEZ"]["matches"] and by_name["JORGE FERNANDEZ"]["matches"][0]["source"] == "pep"
    assert by_name["John Smith"]["row"]["geboortejaar"] == 1971


def test_batch_xlsx_date_cell_birth_year(tmp_path, monkeypatch):
    from datetime import datetime
    from io import BytesIO

    from openpyxl import Workbook

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["Naam", "Geboortejaar"])
    ws.append(["JORGE FERNANDEZ", datetime(1965, 3, 1)])
    ws.append(["John Smith", datetime(1971, 1, 1)])
    buffer = BytesIO()
    wb.save(buffer)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.post("/api/batch", files={"file": ("lijst.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 200, resp.text
    data = _wait_done(client, resp.json()["batch_id"])
    assert data["status"] == "done"
    assert data["errors"] == []
    by_name = {r["row"]["naam"]: r for r in data["rows"]}
    assert by_name["JORGE FERNANDEZ"]["row"]["geboortejaar"] == 1965
    assert by_name["John Smith"]["row"]["geboortejaar"] == 1971
    assert by_name["JORGE FERNANDEZ"]["matches"] and by_name["JORGE FERNANDEZ"]["matches"][0]["source"] == "pep"


def test_batch_xlsx_blank_and_styled_rows_ignored(tmp_path, monkeypatch):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["Naam"])
    ws.append(["JORGE FERNANDEZ"])
    ws.append([None])
    ws.append([None])
    ws["A5"].font = Font(bold=True)
    buffer = BytesIO()
    wb.save(buffer)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.post("/api/batch", files={"file": ("lijst.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 200, resp.text
    data = _wait_done(client, resp.json()["batch_id"])
    assert data["status"] == "done"
    assert data["errors"] == []
    assert data["total"] == 1
    assert [r["row"]["naam"] for r in data["rows"]] == ["JORGE FERNANDEZ"]


def test_batch_corrupt_xlsx_returns_400(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.post("/api/batch", files={"file": ("lijst.xlsx", b"dit is geen excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert resp.status_code == 400
    assert resp.json()["detail"] == "Ongeldig Excel-bestand"


def test_batch_oversized_upload_413_before_parse(tmp_path, monkeypatch):
    import app.main as main

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    monkeypatch.setattr(main.batch, "parse_input", lambda *a, **k: (_ for _ in ()).throw(AssertionError("parse_input mag niet worden aangeroepen")))
    monkeypatch.setattr(main.batch, "MAX_BATCH_BYTES", 10)
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.post("/api/batch", files={"file": ("lijst.csv", b"naam\nJORGE FERNANDEZ\n", "text/csv")})
    assert resp.status_code == 413


def test_batch_empty_file_400(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.post("/api/batch", files={"file": ("lijst.csv", b"", "text/csv")})
    assert resp.status_code == 400


def test_batch_missing_name_column_400(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    resp = client.post("/api/batch", files={"file": ("lijst.csv", _batch_csv("geboortejaar\n1970\n"), "text/csv")})
    assert resp.status_code == 400


def test_batch_report_pdf(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    _wait_done(client, batch_id)
    resp = client.get(f"/api/batch/{batch_id}/report.pdf")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers["content-disposition"]
    assert b".pdf" in resp.headers["content-disposition"].encode()
    assert resp.content[:4] == b"%PDF"


def test_batch_report_csv(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    _wait_done(client, batch_id)
    resp = client.get(f"/api/batch/{batch_id}/report.csv")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "text/csv; charset=utf-8"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:3] == b"\xef\xbb\xbf"
    assert resp.content[3:6] != b"\xef\xbb\xbf"
    body = resp.content.decode("utf-8-sig")
    assert not body.startswith("\ufeff")
    assert body.startswith("regel;")
    assert "JORGE FERNÁNDEZ" in body


def test_batch_report_not_done_or_unknown_404(tmp_path, monkeypatch):
    import app.main as main

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    monkeypatch.setattr(main.batch, "process_job", lambda *a, **k: None)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    assert client.get(f"/api/batch/{batch_id}/report.pdf").status_code == 404
    assert client.get(f"/api/batch/{batch_id}/report.csv").status_code == 404
    assert client.get("/api/batch/onbekend/report.pdf").status_code == 404
    assert client.get("/api/batch/onbekend/report.csv").status_code == 404


def test_batch_creation_logs_audit(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    event = _last_audit_events(tmp_path)[0]
    assert event["method"] == "POST"
    assert event["path"] == "/api/batch"
    assert event["query"]["batch_id"] == batch_id
    assert event["result_count"] == 1


def test_batch_report_logs_audit(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\n")
    _wait_done(client, batch_id)
    client.get(f"/api/batch/{batch_id}/report.pdf")
    events = _last_audit_events(tmp_path)
    assert any(e["path"] == f"/api/batch/{batch_id}/report.pdf" for e in events)


def test_gating_viewer_cannot_batch(monkeypatch, auth_env):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    _create_local_user(role="viewer")
    client = TestClient(create_app(entities=ENTITIES))
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    resp = client.post("/api/batch", files={"file": ("lijst.csv", _batch_csv("naam\nJan\n"), "text/csv")})
    assert resp.status_code == 403


def test_batch_skips_opensanctions_while_search_uses_it(tmp_path, monkeypatch):
    import app.main as main

    calls = []

    def recording_match(*args, **kwargs):
        calls.append(args)
        return []

    monkeypatch.setattr(main.opensanctions, "match_opensanctions", recording_match)
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(
        entities=ENTITIES,
        eu_root=tmp_path,
        pep_root=tmp_path,
        search_db=tmp_path / "search.sqlite",
        os_api_key="KEY",
    ))
    client.get("/api/search", params={"name": "JORGE FERNANDEZ"})
    assert len(calls) == 1
    batch_id = _create_batch(client, "naam\nJORGE FERNANDEZ\nJohn Smith\n")
    _wait_done(client, batch_id)
    assert len(calls) == 1
    assert calls[0][0] == "KEY"


def test_startup_sweep_marks_orphaned_batch_jobs_error(tmp_path, monkeypatch):
    import app.batch as batch_module

    db_path = tmp_path / "batch.sqlite"
    job_id = batch_module.create_job(db_path, "lijst.csv", [{
        "naam": "Jan", "geboortejaar": None, "nationaliteit": None, "geboorteplaats": None, "type": None,
    }])
    TestClient(create_app(entities=ENTITIES))
    job = batch_module.get_job(db_path, job_id)
    assert job["status"] == "error"
    assert job["error_text"] == "Onderbroken door herstart"
    assert job["finished_at"] is not None


def test_watchlist_create_and_list_persist_via_cookie(tmp_path):
    client = TestClient(create_app(entities=ENTITIES))
    resp = client.post("/api/watchlists", json={"label": "Mijn lijst"})
    assert resp.status_code == 200
    wl = resp.json()["watchlist"]
    assert wl["label"] == "Mijn lijst"
    assert wl["id"]
    assert client.cookies.get("watch_key")
    listed = client.get("/api/watchlists").json()["watchlists"]
    assert len(listed) == 1
    assert listed[0]["id"] == wl["id"]
    key = client.cookies.get("watch_key")
    client.get("/api/watchlists")
    assert client.cookies.get("watch_key") == key
    assert len(client.get("/api/watchlists").json()["watchlists"]) == 1


def test_watchlist_rescan_produces_hits_and_dedups(tmp_path, monkeypatch):
    from app import matcher

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={"label": "PEP"}).json()["watchlist"]
    first = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "JORGE FERNANDEZ"})
    assert first.status_code == 200
    data = first.json()
    assert data["watchlist_id"] == wl["id"]
    assert data["new"] == 1
    assert len(data["hits"]) == 1
    hit = data["hits"][0]["match"]
    assert hit["bron"] == "pep"
    assert hit["naam"] == "JORGE FERNÁNDEZ"
    assert hit["datasets"] == ["ar_parliament"]
    assert hit["id"] == "NK-x"
    assert hit["score"] >= matcher.THRESHOLD
    second = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "JORGE FERNANDEZ"})
    assert second.status_code == 200
    assert second.json()["new"] == 0
    assert len(client.get("/api/watchlists/hits").json()["hits"]) == 1


def test_watchlist_hits_endpoint_filters_and_audits(tmp_path, monkeypatch):
    import json

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={"label": "PEP"}).json()["watchlist"]
    client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "JORGE FERNANDEZ"})
    hits = client.get("/api/watchlists/hits").json()["hits"]
    assert len(hits) == 1
    assert hits[0]["watchlist_id"] == wl["id"]
    assert hits[0]["match"]["bron"] == "pep"
    assert len(client.get("/api/watchlists/hits", params={"watchlist_id": wl["id"]}).json()["hits"]) == 1
    assert client.get("/api/watchlists/hits", params={"watchlist_id": "nope"}).json()["hits"] == []
    events = _last_audit_events(tmp_path)
    rescan = [e for e in events if e["path"] == f"/api/watchlists/{wl['id']}/rescan"][0]
    assert rescan["result_count"] == 1
    assert "JORGE FERNANDEZ" not in json.dumps(rescan["query"])
    create = [e for e in events if e["path"] == "/api/watchlists"][0]
    assert create["query"]["action"] == "create"
    assert create["query"]["watchlist_id"] == wl["id"]


def test_watchlist_owner_isolation(tmp_path):
    client_a = TestClient(create_app(entities=ENTITIES))
    client_b = TestClient(create_app(entities=ENTITIES))
    wl = client_a.post("/api/watchlists", json={"label": "geheim van A"}).json()["watchlist"]
    assert client_b.get("/api/watchlists").json()["watchlists"] == []
    assert client_b.get("/api/watchlists/hits").json()["hits"] == []
    assert client_b.delete(f"/api/watchlists/{wl['id']}").status_code == 404
    assert client_b.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "Jan"}).status_code == 404
    assert len(client_a.get("/api/watchlists").json()["watchlists"]) == 1
    assert client_a.delete(f"/api/watchlists/{wl['id']}").status_code == 204
    assert client_a.get("/api/watchlists").json()["watchlists"] == []


def test_watchlist_unknown_and_blank_name_errors(tmp_path):
    client = TestClient(create_app(entities=ENTITIES))
    assert client.delete("/api/watchlists/nope").status_code == 404
    assert client.post("/api/watchlists/nope/rescan", json={"name": "Jan"}).status_code == 404
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    assert wl["label"] == ""
    assert client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "   "}).status_code == 422
    assert client.post(f"/api/watchlists/{wl['id']}/rescan", json={}).status_code == 422


def test_watchlist_need_to_know_watched_name_not_stored(tmp_path, monkeypatch):
    import sqlite3

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={"label": "PEP"}).json()["watchlist"]
    watched = "JORGE FERNANDEZ"
    client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": watched})
    conn = sqlite3.connect(tmp_path / "watchlists.sqlite")
    try:
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'") if not r[0].startswith("sqlite_")]
        assert tables == ["watchlists", "watchlist_hits"]
        for table in tables:
            for row in conn.execute(f"SELECT * FROM {table}"):
                for value in row:
                    if isinstance(value, str):
                        assert watched not in value
    finally:
        conn.close()


def test_status_data_version_changes_when_data_changes(tmp_path, monkeypatch):
    import os
    import time

    import app.main as main

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    monkeypatch.setattr(main.search_index, "rebuild_index", lambda db, eu, pep: main.search_index.build_index(db, [make_eu_entity()], pep))
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    before = client.get("/api/status").json()["data_version"]
    future = time.time() + 5
    os.utime(tmp_path / "ar_parliament" / "entities.ftm.json", (future, future))
    after = client.get("/api/status").json()["data_version"]
    assert before != after


def test_watchlist_gating_requires_login_when_required(auth_env, monkeypatch):
    monkeypatch.setenv("AUTH_REQUIRED", "1")
    _create_local_user(role="viewer")
    client = TestClient(create_app(entities=ENTITIES))
    assert client.get("/api/watchlists").status_code == 401
    client.post("/api/auth/login", json={"username": "alice", "password": "geheim"})
    assert client.get("/api/watchlists").status_code == 200
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    assert client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "Jan"}).status_code == 200


def test_watchlist_client_payload_empty_body_label_blank(tmp_path):
    import sqlite3

    client = TestClient(create_app(entities=ENTITIES))
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    assert wl["label"] == ""
    conn = sqlite3.connect(tmp_path / "watchlists.sqlite")
    try:
        label = conn.execute("SELECT label FROM watchlists WHERE id = ?", (wl["id"],)).fetchone()[0]
    finally:
        conn.close()
    assert label == ""


def test_watchlist_watched_name_absent_after_full_client_cycle(tmp_path, monkeypatch):
    import sqlite3

    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    watched = "JORGE FERNANDEZ"
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    assert wl["label"] == ""
    resp = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": watched, "birth_year": "1965"})
    assert resp.status_code == 200, resp.text
    conn = sqlite3.connect(tmp_path / "watchlists.sqlite")
    try:
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'") if not r[0].startswith("sqlite_")]
        assert tables == ["watchlists", "watchlist_hits"]
        for table in tables:
            for row in conn.execute(f"SELECT * FROM {table}"):
                for value in row:
                    if isinstance(value, str):
                        assert watched not in value
    finally:
        conn.close()


def test_watchlist_rescan_full_criteria_string_birth_year(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    resp = client.post(
        f"/api/watchlists/{wl['id']}/rescan",
        json={
            "name": "JORGE FERNANDEZ",
            "birth_year": "1965",
            "nationality": "AR",
            "birth_place": "Buenos Aires",
            "entity_type": "person",
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["new"] == 1
    hit = data["hits"][0]["match"]
    assert hit["bron"] == "pep"
    assert hit["naam"] == "JORGE FERNÁNDEZ"


def test_watchlist_rescan_invalid_birth_year_422(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    resp = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "JORGE FERNANDEZ", "birth_year": "onbekend"})
    assert resp.status_code == 422
    assert resp.json()["detail"] == "Ongeldig geboortejaar"


def test_to_watchlist_match_eu_naam_never_query_string():
    from app.main import _to_watchlist_match

    result = {
        "source": "eu",
        "score": 90,
        "entity": {"name": "SECRET_QUERY_NAME", "eu_reference_number": "EU.1"},
        "eu": {"matched_alias": None, "total_score": 90},
    }
    match = _to_watchlist_match(result)
    assert match["naam"] != "SECRET_QUERY_NAME"
    assert match["naam"] == "EU.1"


def test_watchlist_eu_hit_naam_is_public_alias_not_query(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_search_db(tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    resp = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "john smith"})
    assert resp.status_code == 200, resp.text
    eu = [h["match"] for h in resp.json()["hits"] if h["match"]["bron"] == "eu"]
    assert eu
    assert eu[0]["naam"] == "John Smith"
    assert eu[0]["naam"] != "john smith"


def test_to_watchlist_match_empty_id_returns_none():
    from app.main import _to_watchlist_match

    result = {
        "source": "pep",
        "score": 95,
        "entity": {"name": "PUBLIC"},
        "pep": {"id": "", "datasets": []},
    }
    assert _to_watchlist_match(result) is None


def test_watchlist_rescan_skips_empty_id_hits(tmp_path, monkeypatch):
    monkeypatch.setenv(search_index.INDEX_ENV, "1")
    _write_pep_fixture(tmp_path)
    entity = make_eu_entity()
    entity["eu_reference_number"] = ""
    search_index.build_index(tmp_path / "search.sqlite", [entity], tmp_path)
    client = TestClient(create_app(entities=ENTITIES, eu_root=tmp_path, pep_root=tmp_path, search_db=tmp_path / "search.sqlite"))
    wl = client.post("/api/watchlists", json={}).json()["watchlist"]
    resp = client.post(f"/api/watchlists/{wl['id']}/rescan", json={"name": "John Smith"})
    assert resp.status_code == 200, resp.text
    assert resp.json()["new"] == 0
    assert client.get("/api/watchlists/hits").json()["hits"] == []
