import sqlite3
import uuid
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import batch


def _csv_bytes(text: str) -> bytes:
    return text.encode("utf-8-sig")


def _xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _row(naam: str, **overrides) -> dict:
    row = {"naam": naam, "geboortejaar": None, "nationaliteit": None, "geboorteplaats": None, "type": None}
    row.update(overrides)
    return row


def test_default_batch_db_default_path(monkeypatch):
    monkeypatch.delenv("BATCH_DB", raising=False)
    assert batch.default_batch_db() == Path("data") / "batch.sqlite"


def test_default_batch_db_env_override(monkeypatch):
    monkeypatch.setenv("BATCH_DB", "/tmp/custom/batch.sqlite")
    assert batch.default_batch_db() == Path("/tmp/custom/batch.sqlite")


def test_init_batch_db_creates_tables(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    batch.init_batch_db(db_path)
    conn = sqlite3.connect(db_path)
    job_columns = [r[1] for r in conn.execute("PRAGMA table_info(batch_jobs)")]
    result_columns = [r[1] for r in conn.execute("PRAGMA table_info(batch_results)")]
    conn.close()
    assert job_columns == ["id", "status", "created_at", "finished_at", "progress", "total", "errors", "error_text"]
    assert result_columns == ["batch_id", "row_index", "row_json", "matches_json"]


def test_init_batch_db_idempotent(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    batch.init_batch_db(db_path)
    batch.init_batch_db(db_path)


def test_parse_csv_semicolon_bom_diacritics():
    content = _csv_bytes("Naam;Geboortejaar;Nationaliteit\nBjörk;1965;IJslands\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert rows == [_row("Björk", geboortejaar=1965, nationaliteit="IJslands")]


def test_parse_csv_comma_delimiter():
    content = _csv_bytes("naam,type\nAcme BV,enterprise\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert rows == [_row("Acme BV", type="enterprise")]


def test_parse_csv_header_case_insensitive():
    content = _csv_bytes("NAME;GeboorteJaar\nJan Jansen;1970\n")
    rows, _ = batch.parse_input("lijst.csv", content)
    assert rows[0]["naam"] == "Jan Jansen"
    assert rows[0]["geboortejaar"] == 1970


def test_parse_csv_row_without_name_is_per_row_error():
    content = _csv_bytes("naam;geboortejaar\nJan;1970\n;1980\nPiet;1990\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert [r["naam"] for r in rows] == ["Jan", "Piet"]
    assert errors == [{"row_index": 3, "error": "Ontbrekende naam"}]


def test_parse_csv_empty_file_raises():
    with pytest.raises(batch.BatchInputError):
        batch.parse_input("lijst.csv", b"")


def test_parse_csv_missing_name_column_raises():
    content = _csv_bytes("geboortejaar;nationaliteit\n1970;NL\n")
    with pytest.raises(batch.BatchInputError):
        batch.parse_input("lijst.csv", content)


def test_parse_csv_quoted_field_with_embedded_newline():
    content = _csv_bytes('naam;geboortejaar\n"Van der\nBerg";1970\n"De Vries";1980\n')
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert rows == [
        _row("Van der\nBerg", geboortejaar=1970),
        _row("De Vries", geboortejaar=1980),
    ]


def test_parse_csv_leading_blank_line_before_header():
    content = _csv_bytes("\nnaam;geboortejaar\nJan;1970\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert rows == [_row("Jan", geboortejaar=1970)]


def test_parse_csv_cp1252_encoding():
    content = "naam;geboortejaar\nBjörk;1965\n".encode("cp1252")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert rows == [_row("Björk", geboortejaar=1965)]


def test_parse_xlsx():
    content = _xlsx_bytes([["Naam", "Geboortejaar", "Nationaliteit"], ["Björk", 1965, "IS"]])
    rows, errors = batch.parse_input("lijst.xlsx", content)
    assert errors == []
    assert rows == [_row("Björk", geboortejaar=1965, nationaliteit="IS")]


def test_parse_xlsx_row_without_name_is_per_row_error():
    content = _xlsx_bytes([["Naam", "Geboortejaar"], ["Jan", 1970], [None, 1980]])
    rows, errors = batch.parse_input("lijst.xlsx", content)
    assert rows == [_row("Jan", geboortejaar=1970)]
    assert errors == [{"row_index": 3, "error": "Ontbrekende naam"}]


def test_parse_csv_invalid_birth_year_is_per_row_error():
    content = _csv_bytes("naam;geboortejaar\nJan Jansen;1970\nPiet;onbekend\nMarie;1965\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == [{"row_index": 3, "error": "Ongeldig geboortejaar"}]
    assert [r["naam"] for r in rows] == ["Jan Jansen", "Piet", "Marie"]
    assert rows[0]["geboortejaar"] == 1970
    assert rows[1]["geboortejaar"] is None
    assert rows[2]["geboortejaar"] == 1965


def test_parse_xlsx_date_cell_birth_year_becomes_year():
    from datetime import datetime

    content = _xlsx_bytes([["Naam", "Geboortejaar"], ["Björk", datetime(1965, 1, 1)]])
    rows, errors = batch.parse_input("lijst.xlsx", content)
    assert errors == []
    assert rows == [_row("Björk", geboortejaar=1965)]


def test_parse_xlsx_corrupt_is_batch_input_error():
    with pytest.raises(batch.BatchInputError) as exc_info:
        batch.parse_input("lijst.xlsx", b"dit is geen excel")
    assert exc_info.value.status_code == 400
    assert str(exc_info.value) == "Ongeldig Excel-bestand"


def test_parse_xlsx_skips_blank_and_styled_rows():
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.append(["Naam", "Geboortejaar"])
    ws.append(["Jan", 1970])
    ws.append([None, None])
    ws.append(["Piet", 1980])
    ws["A6"].font = Font(bold=True)
    buffer = BytesIO()
    wb.save(buffer)
    rows, errors = batch.parse_input("lijst.xlsx", buffer.getvalue())
    assert errors == []
    assert [r["naam"] for r in rows] == ["Jan", "Piet"]
    assert [r["geboortejaar"] for r in rows] == [1970, 1980]


def test_parse_xlsx_blank_rows_do_not_count_toward_limit():
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.append(["Naam"])
    ws.append(["Jan"])
    ws.append([None])
    ws["A5"].font = Font(bold=True)
    buffer = BytesIO()
    wb.save(buffer)
    rows, errors = batch.parse_input("lijst.xlsx", buffer.getvalue(), row_limit=1)
    assert errors == []
    assert [r["naam"] for r in rows] == ["Jan"]


def test_parse_input_over_row_limit_signals_413():
    content = _csv_bytes("naam\n" + "\n".join(f"persoon-{i}" for i in range(5001)) + "\n")
    with pytest.raises(batch.RowLimitExceeded) as exc_info:
        batch.parse_input("lijst.csv", content)
    assert exc_info.value.status_code == 413


def test_parse_input_at_row_limit_ok():
    content = _csv_bytes("naam\n" + "\n".join(f"persoon-{i}" for i in range(5000)) + "\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == []
    assert len(rows) == 5000


def test_create_job_returns_uuid_and_roundtrips(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [_row("Jan"), _row("Acme BV", type="enterprise")]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    uuid.UUID(job_id)
    job = batch.get_job(db_path, job_id)
    assert job["id"] == job_id
    assert job["status"] == "pending"
    assert job["total"] == 2
    assert job["progress"] == 0
    assert job["errors"] == []
    assert job["finished_at"] is None
    assert job["created_at"]


def test_get_job_unknown_returns_none(tmp_path):
    assert batch.get_job(tmp_path / "batch.sqlite", "onbekend") is None


def test_create_job_stores_per_row_errors(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    job_id = batch.create_job(
        db_path,
        "lijst.csv",
        [_row("Jan")],
        errors=[{"row_index": 3, "error": "Ontbrekende naam"}],
    )
    job = batch.get_job(db_path, job_id)
    assert job["errors"] == [{"row_index": 3, "error": "Ontbrekende naam"}]


def test_create_job_roundtrips_parse_input_errors(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    content = _csv_bytes("naam;geboortejaar\nJan;1970\n;1980\n")
    rows, errors = batch.parse_input("lijst.csv", content)
    assert errors == [{"row_index": 3, "error": "Ontbrekende naam"}]
    job_id = batch.create_job(db_path, "lijst.csv", rows, errors=errors)
    job = batch.get_job(db_path, job_id)
    assert job["errors"] == [{"row_index": 3, "error": "Ontbrekende naam"}]
    assert job["total"] == 1


def test_get_results_roundtrip(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [_row("Jan", geboortejaar=1970), _row("Acme BV", type="enterprise")]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    results = batch.get_results(db_path, job_id)
    assert [r["row"] for r in results] == rows
    assert [r["row_index"] for r in results] == [0, 1]
    assert all(r["matches"] is None for r in results)


def test_process_job_fake_search_fn(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [
        _row("Jan", geboortejaar=1970, nationaliteit="NL", geboorteplaats="Amsterdam", type="person"),
        _row("Acme BV", type="enterprise"),
    ]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    calls = []

    def fake_search_fn(naam, geboortejaar=None, nationaliteit=None, geboorteplaats=None, type=None):
        calls.append({"naam": naam, "geboortejaar": geboortejaar, "nationaliteit": nationaliteit, "geboorteplaats": geboorteplaats, "type": type})
        return [{"matched": naam}]

    batch.process_job(db_path, job_id, fake_search_fn)
    job = batch.get_job(db_path, job_id)
    assert job["status"] == "done"
    assert job["progress"] == 2
    assert job["finished_at"] is not None
    assert calls == [
        {"naam": "Jan", "geboortejaar": 1970, "nationaliteit": "NL", "geboorteplaats": "Amsterdam", "type": "person"},
        {"naam": "Acme BV", "geboortejaar": None, "nationaliteit": None, "geboorteplaats": None, "type": "enterprise"},
    ]
    results = batch.get_results(db_path, job_id)
    assert [r["matches"] for r in results] == [[{"matched": "Jan"}], [{"matched": "Acme BV"}]]


def test_process_job_unknown_job_raises(tmp_path):
    with pytest.raises(KeyError):
        batch.process_job(tmp_path / "batch.sqlite", "onbekend", lambda naam, **kw: [])


def test_process_job_row_limit_signals_413(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [_row(f"persoon-{i}") for i in range(5001)]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    with pytest.raises(batch.RowLimitExceeded) as exc_info:
        batch.process_job(db_path, job_id, lambda naam, **kw: [])
    assert exc_info.value.status_code == 413


def test_process_job_twice_is_idempotent(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [_row("Jan"), _row("Piet")]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    batch.process_job(db_path, job_id, lambda naam, **kw: [])
    batch.process_job(db_path, job_id, lambda naam, **kw: [])
    job = batch.get_job(db_path, job_id)
    assert job["status"] == "done"
    assert job["progress"] == 2


def test_process_job_resume_after_error_reports_total_progress(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    rows = [_row(f"p-{i}") for i in range(5)]
    job_id = batch.create_job(db_path, "lijst.csv", rows)
    calls = {"count": 0}

    def failing_search_fn(naam, **kwargs):
        calls["count"] += 1
        if calls["count"] == 3:
            raise RuntimeError("boom")
        return []

    with pytest.raises(RuntimeError):
        batch.process_job(db_path, job_id, failing_search_fn)
    job = batch.get_job(db_path, job_id)
    assert job["status"] == "error"
    assert job["progress"] == 2

    batch.process_job(db_path, job_id, lambda naam, **kwargs: [{"matched": naam}])
    job = batch.get_job(db_path, job_id)
    assert job["status"] == "done"
    assert job["progress"] == 5


def test_mark_stale_jobs_flags_pending_and_running(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    pending_id = batch.create_job(db_path, "lijst.csv", [_row("Jan")])
    running_id = batch.create_job(db_path, "lijst.csv", [_row("Piet")])
    done_id = batch.create_job(db_path, "lijst.csv", [_row("Kees")])
    with batch._open(db_path) as conn:
        conn.execute("UPDATE batch_jobs SET status = 'running' WHERE id = ?", (running_id,))
        conn.execute("UPDATE batch_jobs SET status = 'done' WHERE id = ?", (done_id,))
        conn.commit()
    updated = batch.mark_stale_jobs(db_path)
    assert updated == 2
    assert batch.get_job(db_path, pending_id)["status"] == "error"
    assert batch.get_job(db_path, pending_id)["error_text"] == "Onderbroken door herstart"
    assert batch.get_job(db_path, pending_id)["finished_at"] is not None
    assert batch.get_job(db_path, running_id)["status"] == "error"
    assert batch.get_job(db_path, done_id)["status"] == "done"
    assert batch.get_job(db_path, done_id)["error_text"] is None


def test_mark_stale_jobs_custom_statuses_and_text(tmp_path):
    db_path = tmp_path / "batch.sqlite"
    pending_id = batch.create_job(db_path, "lijst.csv", [_row("Jan")])
    done_id = batch.create_job(db_path, "lijst.csv", [_row("Piet")])
    with batch._open(db_path) as conn:
        conn.execute("UPDATE batch_jobs SET status = 'done' WHERE id = ?", (done_id,))
        conn.commit()
    updated = batch.mark_stale_jobs(db_path, stale_statuses=("done",), error_text="zelf")
    assert updated == 1
    assert batch.get_job(db_path, done_id)["status"] == "error"
    assert batch.get_job(db_path, done_id)["error_text"] == "zelf"
    assert batch.get_job(db_path, pending_id)["status"] == "pending"


def test_mark_stale_jobs_missing_db_is_noop(tmp_path):
    assert batch.mark_stale_jobs(tmp_path / "batch.sqlite") == 0
